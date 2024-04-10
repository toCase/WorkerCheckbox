# This Python file uses the following encoding: utf-8
from PySide6.QtCore import QObject, Slot
from openpyxl import load_workbook
import requests
import json
import uuid
import platform

from Base import Base

class XLS_Maker(QObject):



    def __init__(self):
        QObject.__init__(self)
        Base.connectDB(self)

    @Slot(str, result=str)
    def url_name(self, file :str):
        os_name = platform.system()
        print(os_name)
        if os_name == "Linux":
            # file = "/home/ev/Документи/rro_test.xlsx"
            file = "/home/ev/Документи/test.xlsx"
        else:
            file = file.replace("file:///", "")
            file = file.replace("/", "\\")
        return file

    @Slot(str, result=list)
    def get_data(self, fname:str):
        print("DEL DEBUG DATA")
        Base.order_del_debug(self)

        print("GET DATA")

        print("File: ", fname)


        if fname != "":

            wb = load_workbook(fname)
            sheets = wb.sheetnames
            sheet = wb[sheets[0]]

            row_start = 2
            row_count = 1
            for row in sheet.iter_rows(values_only=True):
                row_count = row_count + 1

                # test_cell = sheet["B" + str(row)].value
                # print("TS: ", test_cell)

            # print("RC:",row_count)

            for r in range(row_start, row_count, 1):
                phone = "38" + str(sheet["A" + str(r)].value).replace("-", "").replace(" ", "")
                product = sheet["B" + str(r)].value

                # print("PRODUCT", product)

                if product != None:
                
                    code = product.split(" ")[-1]
                    # print("CODE: ", code)
                    product_name = Base.product_getName(self, code)

                    quan = str(sheet["C" + str(r)].value)
                    ttn = sheet["D" + str(r)].value

                    product_count = len(product.split("\n"))

                    # print ("Count: ", product_count)

                    #make order
                    otype = 1 if ttn == None else 2

                    l = {
                        "id":0,
                        "type":otype,
                        "phone":phone,
                        "ttn":ttn,
                        "debug":1
                    }

                    order = Base.order_save(self, l)
                    order_id = order.get('id')

                    if order.get('r') == True:
                        if product_count == 0:

                            d = {
                                "id":0,
                                "p_order":order_id,
                                "p_name":product_name,
                                "p_code":code,
                                "p_quan":quan.split(" ")[0],
                                "p_ids":Base.product_getIDS(self, code),
                            }

                            detail = Base.detail_save(d)

                            if detail.get('r') == False:
                                return [0, detail.get("error")]

                        elif product_count > 0:
                            for p in range(0, product_count, 1):

                                prod = product.split("\n")[p]

                                # print("PROD:", prod)
                                # print(len(str(prod).split(" ")))
                                product_code = str(prod).split(" ")[-1]

                                product_name = Base.product_getName(self, product_code)

                                d = {
                                    "id":0,
                                    "p_order":order_id,
                                    "p_name":product_name,
                                    "p_code":product_code,
                                    "p_quan":quan.split("\n")[p].split(" ")[0],
                                    "p_ids":Base.product_getIDS(self, code),
                                }

                                detail = Base.detail_save(d)
                                if detail.get('r') == False:
                                    return [0, detail.get("error")]
                    else:
                        return [0, order.get("error")]
            return [1, ""]
        else:
            return [0, "No file"]

